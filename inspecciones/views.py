import os
import io
import requests
from datetime import date
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.conf import settings
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from PIL import Image as PILImage

from .models import RegistroInspeccion, Medidor
from .forms import RegistroForm, MedidorFormSet

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, Image
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas

# 🔹 Fuente
VERDANA_PATH = r'C:\Windows\Fonts\verdana.ttf'
if os.path.exists(VERDANA_PATH):
    pdfmetrics.registerFont(TTFont('Verdana', VERDANA_PATH))
    FONT = 'Verdana'
else:
    FONT = 'Helvetica'


# ─────────────────────────────────────────────────────────────
# 🔹 INDEX → redirige directo al importador
# ─────────────────────────────────────────────────────────────
@login_required
def index(request):
    return redirect('importar_excel')


# ─────────────────────────────────────────────────────────────
# 🔹 DETALLE
# ─────────────────────────────────────────────────────────────
@login_required
def detalle_registro(request, pk):
    registro = get_object_or_404(RegistroInspeccion, pk=pk)
    return render(request, 'detalle.html', {'registro': registro})


# ─────────────────────────────────────────────────────────────
# 🔹 CREAR (se mantiene por compatibilidad)
# ─────────────────────────────────────────────────────────────
@login_required
def nuevo_registro(request):
    if request.method == 'POST':
        form = RegistroForm(request.POST, request.FILES)
        formset = MedidorFormSet(request.POST, request.FILES)
        if form.is_valid() and formset.is_valid():
            registro = form.save()
            medidores = formset.save(commit=False)
            for m in medidores:
                m.registro = registro
                m.save()
            from django.contrib import messages
            messages.success(request, '✅ Registro guardado correctamente.')
            return redirect('importar_excel')
        else:
            print("FORM ERRORS:", form.errors)
            print("FORMSET ERRORS:", formset.errors)
    else:
        form = RegistroForm()
        formset = MedidorFormSet()
    return render(request, 'nuevo.html', {'form': form, 'formset': formset})


# ─────────────────────────────────────────────────────────────
# 🔹 EDITAR
# ─────────────────────────────────────────────────────────────
@login_required
def editar_registro(request, pk):
    registro = get_object_or_404(RegistroInspeccion, pk=pk)
    if request.method == 'POST':
        form = RegistroForm(request.POST, request.FILES, instance=registro)
        formset = MedidorFormSet(request.POST, request.FILES, instance=registro)
        if form.is_valid() and formset.is_valid():
            registro = form.save()
            medidores = formset.save(commit=False)
            for m in medidores:
                m.registro = registro
                m.save()
            for obj in formset.deleted_objects:
                obj.delete()
            return redirect('importar_excel')
    else:
        form = RegistroForm(instance=registro)
        formset = MedidorFormSet(instance=registro)
    return render(request, 'nuevo.html', {'form': form, 'formset': formset})


# ─────────────────────────────────────────────────────────────
# 🔹 ELIMINAR
# ─────────────────────────────────────────────────────────────
@login_required
def eliminar_registro(request, pk):
    registro = get_object_or_404(RegistroInspeccion, pk=pk)
    if request.method == 'POST':
        registro.delete()
    return redirect('importar_excel')


# ─────────────────────────────────────────────────────────────
# 🔹 EXPORTAR EXCEL
# ─────────────────────────────────────────────────────────────
@login_required
def exportar_excel(request):
    fecha_filtro = request.GET.get('fecha', '')
    if fecha_filtro:
        registros = RegistroInspeccion.objects.filter(fecha_informe=fecha_filtro)
    else:
        registros = RegistroInspeccion.objects.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Registros'

    encabezado = ['Fecha Informe', 'Fecha Despiece', 'Serial', 'Modelo', 'Año',
                  'Estado', 'Código', 'Alteración', 'Observaciones']
    for col, titulo in enumerate(encabezado, 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    fila = 2
    for r in registros:
        for m in r.medidores.all():
            ws.cell(row=fila, column=1, value=str(r.fecha_informe))
            ws.cell(row=fila, column=2, value=str(r.fecha_despiece))
            ws.cell(row=fila, column=3, value=m.serial)
            ws.cell(row=fila, column=4, value=m.modelo)
            ws.cell(row=fila, column=5, value=m.anio)
            ws.cell(row=fila, column=6, value=m.estado)
            ws.cell(row=fila, column=7, value=m.codigo)
            ws.cell(row=fila, column=8, value=m.medidor_con_alteracion)
            ws.cell(row=fila, column=9, value=m.observaciones_encontradas)
            fila += 1

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="registros.xlsx"'
    wb.save(response)
    return response


# ─────────────────────────────────────────────────────────────
# 🔹 SELECCIONAR MEDIDORES (BD — se mantiene)
# ─────────────────────────────────────────────────────────────
@login_required
def seleccionar_medidores(request):
    fecha = request.GET.get('fecha')
    if fecha:
        medidores = Medidor.objects.filter(registro__fecha_informe=fecha)
    else:
        medidores = Medidor.objects.all()
    return render(request, 'seleccionar.html', {'medidores': medidores, 'fecha': fecha})


# ─────────────────────────────────────────────────────────────
# 🔹 CANVAS CON PAGINACIÓN
# ─────────────────────────────────────────────────────────────
class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        total = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_footer(total)
            super().showPage()
        super().save()

    def draw_footer(self, total):
        self.setFont(FONT, 8)
        self.drawCentredString(
            300, 20,
            f"Página {self._pageNumber} de {total} / MPE-02-F-28-01 Versión: 12 / 2025-08-01"
        )


# ─────────────────────────────────────────────────────────────
# 🔹 HELPERS COMPARTIDOS PDF
# ─────────────────────────────────────────────────────────────
def _get_styles():
    styles = getSampleStyleSheet()
    styles['Normal'].fontName = FONT
    styles['Normal'].fontSize = 9
    styles['Normal'].alignment = 4
    styles['Heading2'].fontName = FONT
    styles['Heading2'].fontSize = 10
    styles['Heading3'].fontName = FONT
    styles['Heading3'].fontSize = 9

    titulo_style = ParagraphStyle(
        'TituloInforme', fontName=FONT, fontSize=11,
        alignment=TA_CENTER, spaceAfter=10,
    )
    firma_style = ParagraphStyle(
        'FirmaStyle', fontName=FONT, fontSize=9, alignment=TA_CENTER,
    )
    return styles, titulo_style, firma_style


def _header_footer(canvas_obj, doc):
    canvas_obj.saveState()
    try:
        logo = os.path.join(settings.MEDIA_ROOT, 'logo.png')
        canvas_obj.drawImage(logo, 40, 750, width=520, height=70)
    except Exception:
        pass
    try:
        pie = os.path.join(settings.MEDIA_ROOT, 'pie.png')
        canvas_obj.drawImage(pie, 40, 30, width=520, height=60)
    except Exception:
        pass
    canvas_obj.restoreState()


def _build_pdf_elements(medidores_data, fecha_informe_str, fecha_despiece_str,
                        styles, titulo_style, firma_style):
    """
    Construye los elementos ReportLab del informe.
    medidores_data: lista de dicts con claves:
        serial, modelo, anio, estado, codigo, alteracion, observaciones,
        fotos (lista de hasta 4 Image de ReportLab o "")
    """
    elementos = []

    elementos.append(Paragraph(
        "<b>INFORME TÉCNICO DE EVALUACIÓN DESPIECE DE MEDIDORES</b>",
        titulo_style
    ))
    elementos.append(Spacer(1, 10))
    elementos.append(Paragraph(
        f"<b>Fecha del informe:</b> {fecha_informe_str}", styles['Normal']))
    elementos.append(Paragraph(
        f"<b>Fecha del despiece:</b> {fecha_despiece_str}", styles['Normal']))
    elementos.append(Spacer(1, 12))

    # ── Sección 1: tabla resumen ──────────────────────────────
    elementos.append(Paragraph("<b>1. INFORMACIÓN DEL MEDIDOR</b>", styles['Heading2']))
    header = ['SERIAL', 'MODELO', 'AÑO', 'ESTADO', 'CODIGO', 'MEDIDOR CON ALTERACIÓN']
    data_tabla = [header]
    for m in medidores_data:
        data_tabla.append([
            m['serial'], m['modelo'], str(m['anio']),
            m['estado'], m['codigo'], m['alteracion']
        ])

    tabla = Table(data_tabla, colWidths=[75, 75, 35, 65, 65, 130])
    tabla_style = [
        ('BACKGROUND',     (0, 0), (-1, 0),  colors.HexColor("#E8E8E8")),
        ('TEXTCOLOR',      (0, 0), (-1, 0),  colors.black),
        ('FONTNAME',       (0, 0), (-1, -1), FONT),
        ('FONTSIZE',       (0, 0), (-1, -1), 8),
        ('ALIGN',          (0, 0), (-1, -1), 'CENTER'),
        ('GRID',           (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor("#C8DDE5"), colors.white]),
        # Columna ALTERACIÓN: fondo rosado por defecto
        ('BACKGROUND',     (5, 0), (5, 0),  colors.HexColor("#E3CFCF")),
        ('BACKGROUND',     (5, 1), (5, -1), colors.HexColor("#E3CFCF")),
    ]

    # Si alteración es SI → fondo rojo y texto blanco
    for i, m in enumerate(medidores_data, 1):
        if str(m['alteracion']).upper() == 'SI':
            tabla_style.append(('BACKGROUND', (5, i), (5, i), colors.red))
            tabla_style.append(('TEXTCOLOR',  (5, i), (5, i), colors.white))

    tabla.setStyle(tabla_style)
    elementos.append(tabla)
    elementos.append(Spacer(1, 12))

    # ── Sección 2: observaciones + fotos por medidor ──────────
    elementos.append(Paragraph("<b>2. OBSERVACIÓN DEL DESPIECE</b>", styles['Heading2']))

    for m in medidores_data:
        elementos.append(Spacer(1, 8))
        elementos.append(Paragraph(
            f"<b>Serial:</b> {m['serial']} &nbsp;&nbsp; <b>Modelo:</b> {m['modelo']}",
            styles['Heading3']
        ))
        elementos.append(Paragraph("<b>Observación encontrada</b>", styles['Heading3']))
        if m['observaciones']:
            elementos.append(Paragraph(m['observaciones'], styles['Normal']))
        elementos.append(Spacer(1, 6))

        fotos = list(m.get('fotos', []))
        while len(fotos) < 4:
            fotos.append("")

        tabla_fotos = Table(
            [[fotos[0], fotos[1]], [fotos[2], fotos[3]]],
            colWidths=[130, 130]
        )
        tabla_fotos.setStyle([
            ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING',    (0, 0), (-1, -1), 5),
            ('LEFTPADDING',   (0, 0), (-1, -1), 5),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 5),
        ])
        elementos.append(tabla_fotos)

    elementos.append(Spacer(1, 20))

    # ── Firma ─────────────────────────────────────────────────
    firma_path = os.path.join(settings.MEDIA_ROOT, 'firma.png')
    if os.path.exists(firma_path):
        firma_tabla_data = [[Image(firma_path, width=100, height=50)]]
    else:
        firma_tabla_data = [['']]

    firma_tabla = Table(firma_tabla_data, colWidths=[500])
    firma_tabla.setStyle([('ALIGN', (0, 0), (0, 0), 'CENTER')])
    elementos.append(firma_tabla)
    elementos.append(Paragraph("<b>PAULA JULIA BLANCO H</b>", firma_style))
    elementos.append(Paragraph(
        "Líder CN Laboratorio de Calibración de Medidores", firma_style))

    return elementos


# ─────────────────────────────────────────────────────────────
# 🔹 GENERAR INFORME PDF (desde BD — se mantiene)
# ─────────────────────────────────────────────────────────────
@login_required
def generar_informe(request):
    ids = request.POST.getlist('medidores')
    medidores = Medidor.objects.filter(id__in=ids) if ids else Medidor.objects.none()

    fecha_informe = ''
    fecha_despiece = ''
    if medidores.exists():
        registro = medidores.first().registro
        if registro.fecha_informe:
            fecha_informe = registro.fecha_informe.strftime('%Y-%m-%d')
        if registro.fecha_despiece:
            fecha_despiece = registro.fecha_despiece.strftime('%Y-%m-%d')

    styles, titulo_style, firma_style = _get_styles()

    medidores_data = []
    for m in medidores:
        fotos = []
        for foto in [m.foto_1, m.foto_2, m.foto_3, m.foto_4]:
            if foto:
                try:
                    fotos.append(Image(foto.path, width=120, height=90))
                except Exception:
                    fotos.append("")
            else:
                fotos.append("")
        medidores_data.append({
            'serial':        m.serial,
            'modelo':        m.modelo,
            'anio':          str(m.anio),
            'estado':        m.estado,
            'codigo':        m.codigo,
            'alteracion':    m.medidor_con_alteracion,
            'observaciones': m.observaciones_encontradas,
            'fotos':         fotos,
        })

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="informe.pdf"'
    doc = SimpleDocTemplate(response, rightMargin=40, leftMargin=40,
                            topMargin=120, bottomMargin=100)
    elementos = _build_pdf_elements(
        medidores_data, fecha_informe, fecha_despiece,
        styles, titulo_style, firma_style
    )
    doc.build(elementos, onFirstPage=_header_footer,
              onLaterPages=_header_footer, canvasmaker=NumberedCanvas)
    return response


# ─────────────────────────────────────────────────────────────
# 🔹 FLUJO EXCEL: LEER → PREVISUALIZAR → PDF
# ─────────────────────────────────────────────────────────────

def _descargar_foto(url):
    """
    Descarga imagen desde URL de SharePoint y devuelve Image de ReportLab.
    Si falla o la URL está vacía, devuelve "".

    Para autenticación con SharePoint agrega:
        headers = {'Authorization': 'Bearer TU_TOKEN'}
        resp = requests.get(url.strip(), headers=headers, timeout=15)
    """
    if not url or not url.strip().startswith('http'):
        return ""
    try:
        resp = requests.get(url.strip(), timeout=15)
        if resp.status_code == 200:
            img = PILImage.open(io.BytesIO(resp.content)).convert('RGB')
            buf = io.BytesIO()
            img.save(buf, format='JPEG')
            buf.seek(0)
            return Image(buf, width=120, height=90)
    except Exception:
        pass
    return ""


def _leer_excel(archivo):
    """
    Lee el Excel exportado de Microsoft Forms.
    - Año: toma 'Año', si está vacío usa 'Año1'
    - Fecha despiece: columna 'Hora de inicio'
    - Fotos: una celda puede contener varias URLs separadas por ';'
    - Alteración: columna 'MEDIDOR CON ALTERACIÓN'
    Devuelve lista de dicts.
    """
    wb = load_workbook(filename=archivo, read_only=True)
    ws = wb.active

    encabezados = {}
    for cell in next(ws.iter_rows(max_row=1, values_only=False)):
        if cell.value is not None:
            encabezados[str(cell.value).strip()] = cell.column - 1

    def idx(nombre):
        return encabezados.get(nombre)

    medidores = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        def v(nombre):
            i = idx(nombre)
            return str(row[i]).strip() if i is not None and row[i] is not None else ''

        # Año: preferir 'Año', sino 'Año1'
        anio = v('Año') or v('Año1')

        # Fecha despiece desde 'Hora de inicio'
        i_hora = idx('Hora de inicio')
        fecha_despiece = ''
        if i_hora is not None and row[i_hora] is not None:
            val = row[i_hora]
            if hasattr(val, 'strftime'):
                fecha_despiece = val.strftime('%Y-%m-%d')
            else:
                fecha_despiece = str(val)[:10]

        # Fotos: varias URLs posibles por celda separadas por ';'
        urls_fotos = []
        for campo in ['Registro fotográfico 1', 'Registro fotográfico 2',
                      'Registro fotográfico 3', 'Registro fotográfico 4']:
            i_f = idx(campo)
            if i_f is not None and row[i_f]:
                for url in str(row[i_f]).split(';'):
                    url = url.strip()
                    if url:
                        urls_fotos.append(url)

        # Alteración
        alteracion = v('MEDIDOR CON ALTERACIÓN')

        medidores.append({
            'serial':         v('Serial'),
            'modelo':         v('Modelo'),
            'anio':           anio,
            'estado':         v('Estado'),
            'codigo':         v('Codigo'),
            'observaciones':  v('Observaciones'),
            'alteracion':     alteracion,
            'fecha_despiece': fecha_despiece,
            'urls_fotos':     urls_fotos[:4],
        })

    return medidores


@login_required
def importar_excel(request):
    """
    GET  → formulario para subir el Excel.
    POST → lee el Excel, guarda en sesión y redirige a previsualización.
    """
    from django.contrib import messages

    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        archivo = request.FILES['archivo_excel']
        try:
            medidores = _leer_excel(archivo)
        except Exception as e:
            messages.error(request, f'Error al leer el archivo: {e}')
            return redirect('importar_excel')

        if not medidores:
            messages.error(request, 'El archivo no contiene datos válidos.')
            return redirect('importar_excel')

        request.session['medidores_excel'] = medidores
        return redirect('previsualizar_excel')

    return render(request, 'importar_excel.html')


@login_required
def previsualizar_excel(request):
    """
    Muestra tabla con los registros del Excel.
    El usuario selecciona cuáles incluir en el PDF.
    """
    medidores = request.session.get('medidores_excel', [])
    if not medidores:
        return redirect('importar_excel')
    return render(request, 'previsualizar_excel.html', {'medidores': medidores})


@login_required
def generar_pdf_excel(request):
    """
    Recibe índices seleccionados, descarga fotos desde SharePoint
    y genera el PDF con el formato estándar del informe.
    """
    if request.method != 'POST':
        return redirect('importar_excel')

    from django.contrib import messages

    todos = request.session.get('medidores_excel', [])
    indices = request.POST.getlist('indices')

    if not indices:
        messages.error(request, 'Debes seleccionar al menos un registro.')
        return redirect('previsualizar_excel')

    seleccionados = [todos[int(i)] for i in indices if int(i) < len(todos)]

    # Fecha informe = hoy | Fecha despiece = del primer registro
    fecha_informe_str = date.today().strftime('%Y-%m-%d')
    fecha_despiece_str = seleccionados[0].get('fecha_despiece', '') if seleccionados else ''

    styles, titulo_style, firma_style = _get_styles()

    medidores_data = []
    for m in seleccionados:
        fotos = [_descargar_foto(url) for url in m.get('urls_fotos', [])]
        while len(fotos) < 4:
            fotos.append("")
        medidores_data.append({
            'serial':        m['serial'],
            'modelo':        m['modelo'],
            'anio':          m['anio'],
            'estado':        m['estado'],
            'codigo':        m['codigo'],
            'alteracion':    m.get('alteracion', 'NO'),
            'observaciones': m['observaciones'],
            'fotos':         fotos,
        })

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="informe_excel.pdf"'
    doc = SimpleDocTemplate(response, rightMargin=40, leftMargin=40,
                            topMargin=120, bottomMargin=100)
    elementos = _build_pdf_elements(
        medidores_data, fecha_informe_str, fecha_despiece_str,
        styles, titulo_style, firma_style
    )
    doc.build(elementos, onFirstPage=_header_footer,
              onLaterPages=_header_footer, canvasmaker=NumberedCanvas)
    return response